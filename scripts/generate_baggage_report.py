#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成行李延误审核结果 Excel 报告"""

import json
import re
from pathlib import Path
from collections import Counter
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REVIEW_DIR = Path('review_results/baggage_delay')
CLAIMS_DIR = Path('claims_data/行李延误')
from datetime import datetime as _dt
OUT_FILE   = Path(f'review_results/行李延误审核结果报告_{_dt.now().strftime("%m%d_%H%M")}.xlsx')


# ─── 辅助 ────────────────────────────────────────────────────────────────
def _s(v, maxlen=None):
    s = '' if v is None else str(v).strip()
    if s.lower() in ('null', 'none', 'unknown', ''): s = ''
    return s[:maxlen] if maxlen and s else s

def _b(v):
    s = str(v).lower().strip() if v is not None else ''
    if s in ('true', '1', 'yes', '是'): return '是'
    if s in ('false', '0', 'no', '否'): return '否'
    return ''

def map_human(s):
    if '支付' in s or '赔付' in s: return 'approve'
    if '拒赔' in s or '零结' in s: return 'reject'
    return 'other'

def map_ai(r):
    audit = r.get('baggage_delay_audit') or {}
    ar = str(audit.get('audit_result') or '').strip()
    if ar in ('通过', 'approve'): return 'approve'
    if ar in ('拒绝', 'reject'):  return 'reject'
    if ar in ('需补件', 'supplement'): return 'supplement'
    is_add = str(r.get('IsAdditional') or 'N').strip().upper()
    if is_add == 'Y': return 'supplement'
    remark = str(r.get('Remark') or '')
    if '审核通过' in remark: return 'approve'
    if '拒赔' in remark or '拒绝' in remark: return 'reject'
    if '需补件' in remark or '补件' in remark or '转人工' in remark: return 'supplement'
    return 'unknown'

def ai_label(r):
    v = map_ai(r)
    return {'approve': '✅ 通过', 'reject': '❌ 拒赔',
            'supplement': '📋 需补件', 'unknown': '❓ 未知'}[v]

def human_label(s):
    v = map_human(s)
    return {'approve': '✅ 通过', 'reject': '❌ 拒赔', 'other': '— 其他'}[v]

def match_label(h, ai):
    if h == 'other': return '—'
    return '✅ 一致' if h == ai else '❌ 不一致'


# ─── 读取数据 ────────────────────────────────────────────────────────────
# claim_map: forceid -> claim_info dict（按 forceid 字段索引，兜底用）
# folder_map: 文件夹路径 -> claim_info dict（按实际文件夹路径索引，优先用）
claim_map = {}
folder_map = {}
for f in CLAIMS_DIR.rglob('claim_info.json'):
    d = json.loads(f.read_text(encoding='utf-8'))
    fid = str(d.get('forceid') or '').strip()
    if fid:
        claim_map[fid] = d
    # 以文件夹路径（正反斜杠统一）为 key
    folder_key = str(f.parent).replace('\\', '/')
    folder_map[folder_key] = d

rows = []
for rf in sorted(REVIEW_DIR.glob('*_ai_review.json')):
    result = json.loads(rf.read_text(encoding='utf-8'))
    forceid = str(result.get('forceid') or '').strip()
    di = result.get('DebugInfo') or {}

    # 优先用 DebugInfo.claim_folder 路径查找，避免 claim_info 里 forceid 写错的问题
    claim_folder = str(di.get('claim_folder') or '').replace('\\', '/')
    ci = folder_map.get(claim_folder) or claim_map.get(forceid, {})

    di     = result.get('DebugInfo') or {}
    vision = di.get('vision_extract') or {}
    parsed = di.get('ai_parsed') or {}
    avi    = di.get('aviation_lookup') or {}
    pv     = di.get('policy_validity') or {}
    missing = di.get('missing_materials') or []

    # 被保险人信息
    passenger    = _s(ci.get('Insured_And_Policy') or ci.get('Applicant_Name') or
                      ci.get('Passenger_Name') or ci.get('insured_name'))
    case_no      = _s(ci.get('ClaimId') or ci.get('CaseNo') or ci.get('case_no'))
    policy_no    = _s(ci.get('PolicyNo') or ci.get('Policy_No') or ci.get('policy_no'))
    h_raw        = _s(ci.get('Final_Status') or ci.get('final_status'))
    accident_date = _s(ci.get('Date_of_Accident') or ci.get('date_of_accident'))

    # 航班信息
    segs      = vision.get('itinerary_segments') or []
    alt       = vision.get('alternate') or {}

    seg_count = len(segs) if segs else 1
    seg_flights = '; '.join(
        f"段{s.get('segment_no')} {_s(s.get('original_flight_no'))} "
        f"{_s(s.get('original_dep_iata'))}→{_s(s.get('original_arr_iata'))} "
        f"{_s(s.get('original_date'))}"
        for s in segs
    ) if segs else ''

    main_flight = _s(vision.get('flight_no') or parsed.get('flight_no'))
    main_date   = _s(vision.get('flight_date') or parsed.get('flight_date'))
    dep_iata    = _s(vision.get('dep_iata') or parsed.get('dep_iata'))
    arr_iata    = _s(vision.get('arr_iata') or parsed.get('arr_iata'))

    # 联程判断
    is_connecting = '是' if len(segs) > 1 else _b(alt.get('is_connecting_rebooking'))
    missed_conn   = _b(alt.get('is_connecting_missed'))

    # 改签信息
    alt_flight  = _s(alt.get('alt_flight_no'))
    alt_dep     = _s(alt.get('alt_dep'))
    alt_arr     = _s(alt.get('alt_arr'))
    alt_source  = _s(alt.get('alt_source'))
    is_rebooking = '是' if alt_flight else '否'

    # 时间计算
    arr_time     = _s(vision.get('flight_actual_arrival_time') or parsed.get('flight_actual_arrival_time'))
    receipt_time = _s(vision.get('baggage_receipt_time') or parsed.get('baggage_receipt_time'))
    delay_hours  = _s(vision.get('delay_hours') or parsed.get('delay_hours'))

    # 材料
    has_boarding = _b(vision.get('has_boarding_or_ticket') or parsed.get('has_boarding_or_ticket'))
    has_delay_pf = _b(vision.get('has_baggage_delay_proof') or parsed.get('has_baggage_delay_proof'))
    delay_pf_src = _s(vision.get('baggage_delay_proof_source'))
    has_receipt  = _b(vision.get('has_baggage_receipt_time_proof') or parsed.get('has_baggage_receipt_time_proof'))
    has_passport = _b(vision.get('has_passport') or parsed.get('has_passport'))
    has_exit     = _b(vision.get('has_exit_entry_record') or parsed.get('has_exit_entry_record'))
    has_bank     = _b(vision.get('has_bank_card_proof') or parsed.get('has_bank_card_proof'))

    # 飞常准
    avi_success = '是' if avi.get('success') else '否'
    avi_actual  = _s(avi.get('actual_arr') or avi.get('actual_dep'))
    avi_status  = _s(avi.get('status'))

    # 保单
    pol_eff   = _s(str(pv.get('effective_date') or ''))[:10]
    pol_exp   = _s(str(pv.get('expiry_date') or ''))[:10]
    pol_valid = '是' if str(pv.get('is_valid') or '').lower() == 'true' else (
                '否' if pv.get('is_valid') is False else '')

    # AI 结论
    ai_result  = ai_label(result)
    ai_verdict = map_ai(result)
    remark     = _s(result.get('Remark'))

    # 赔付金额
    payout = ''
    audit  = result.get('baggage_delay_audit') or {}
    ps     = audit.get('payout_suggestion') or {}
    if ps and ps.get('amount') is not None:
        payout = f"{ps.get('amount')}元"
    if not payout:
        m = re.search(r'赔付([\d.]+)元', remark)
        if m: payout = f"{m.group(1)}元"

    missing_str = '；'.join(missing) if missing else ''
    h_verdict   = map_human(h_raw)

    rows.append({
        'forceid':       forceid,
        'case_no':       case_no,
        'passenger':     passenger,
        'policy_no':     policy_no,
        'accident_date': accident_date,
        'main_flight':   main_flight,
        'main_date':     main_date,
        'dep_iata':      dep_iata,
        'arr_iata':      arr_iata,
        'seg_count':     seg_count,
        'seg_flights':   seg_flights,
        'is_connecting': is_connecting,
        'missed_conn':   missed_conn,
        'is_rebooking':  is_rebooking,
        'alt_flight':    alt_flight,
        'alt_dep':       alt_dep,
        'alt_arr':       alt_arr,
        'alt_source':    alt_source,
        'arr_time':      arr_time,
        'receipt_time':  receipt_time,
        'delay_hours':   delay_hours,
        'avi_success':   avi_success,
        'avi_actual':    avi_actual,
        'avi_status':    avi_status,
        'has_boarding':  has_boarding,
        'has_delay_pf':  has_delay_pf,
        'delay_pf_src':  delay_pf_src,
        'has_receipt':   has_receipt,
        'has_passport':  has_passport,
        'has_exit':      has_exit,
        'has_bank':      has_bank,
        'pol_eff':       pol_eff,
        'pol_exp':       pol_exp,
        'pol_valid':     pol_valid,
        'h_raw':         h_raw,
        'h_verdict':     human_label(h_raw),
        'ai_result':     ai_result,
        'ai_verdict':    ai_verdict,
        'payout':        payout,
        'remark':        remark,
        'missing_str':   missing_str,
        'match_str':     match_label(h_verdict, ai_verdict),
        '_h_verdict':    h_verdict,
    })

print(f'读取 {len(rows)} 条审核记录')


# ─── Excel 样式 ──────────────────────────────────────────────────────────
C_HEADER_BG = '1F3864'
C_HEADER_FG = 'FFFFFF'
C_GROUP_BG  = 'D6E4F7'
C_GROUP_FG  = '1F3864'
C_MATCH_OK  = 'E2EFDA'
C_MATCH_NO  = 'FCE4D6'
C_SUPP      = 'FFF2CC'
C_ODD       = 'F5F8FF'
C_EVEN      = 'FFFFFF'
C_BORDER    = 'BFBFBF'

thin   = Side(style='thin', color=C_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfill(color): return PatternFill('solid', fgColor=color)
def hfont(color='000000', bold=False, size=10):
    return Font(name='微软雅黑', color=color, bold=bold, size=size)
def halign(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# 列定义：(分组, 列标题, 字段key, 宽度, 对齐)
COLS = [
    # 基本信息
    ('基本信息', '序号',            '__idx__',       5,  'center'),
    ('基本信息', 'Force ID',        'forceid',       22, 'left'),
    ('基本信息', '案件号',          'case_no',       14, 'center'),
    ('基本信息', '被保险人',        'passenger',     12, 'center'),
    ('基本信息', '保单号',          'policy_no',     16, 'center'),
    ('基本信息', '事故日期',        'accident_date', 11, 'center'),
    # 航班信息
    ('航班信息', '主航班号',        'main_flight',   10, 'center'),
    ('航班信息', '航班日期',        'main_date',     12, 'center'),
    ('航班信息', '出发机场',        'dep_iata',       8, 'center'),
    ('航班信息', '到达机场',        'arr_iata',       8, 'center'),
    ('航班信息', '行程段数',        'seg_count',      7, 'center'),
    ('航班信息', '是否联程',        'is_connecting',  7, 'center'),
    ('航班信息', '是否误机',        'missed_conn',    7, 'center'),
    ('航班信息', '各段详情',        'seg_flights',   40, 'left'),
    # 改签信息
    ('改签信息', '是否改签',        'is_rebooking',   7, 'center'),
    ('改签信息', '改签航班号',      'alt_flight',    10, 'center'),
    ('改签信息', '改签出发时间',    'alt_dep',       14, 'center'),
    ('改签信息', '改签到达时间',    'alt_arr',       14, 'center'),
    ('改签信息', '改签凭证来源',    'alt_source',    18, 'left'),
    # 时间计算
    ('时间计算', '航班实际到达',    'arr_time',      15, 'center'),
    ('时间计算', '行李签收时间',    'receipt_time',  15, 'center'),
    ('时间计算', '延误时长(h)',     'delay_hours',    9, 'center'),
    # 飞常准
    ('飞常准数据', '飞常准命中',    'avi_success',    8, 'center'),
    ('飞常准数据', '飞常准实际时间','avi_actual',    15, 'center'),
    ('飞常准数据', '飞常准状态',    'avi_status',     9, 'center'),
    # 材料
    ('材料完整性', '登机牌/机票',   'has_boarding',   8, 'center'),
    ('材料完整性', '行李延误证明',  'has_delay_pf',   8, 'center'),
    ('材料完整性', '延误证明来源',  'delay_pf_src',  26, 'left'),
    ('材料完整性', '行李签收证明',  'has_receipt',    8, 'center'),
    ('材料完整性', '护照',          'has_passport',   6, 'center'),
    ('材料完整性', '出入境记录',    'has_exit',       8, 'center'),
    ('材料完整性', '银行卡',        'has_bank',       6, 'center'),
    # 保单
    ('保单信息', '保单生效日',      'pol_eff',       11, 'center'),
    ('保单信息', '保单到期日',      'pol_exp',       11, 'center'),
    ('保单信息', '保单有效',        'pol_valid',      7, 'center'),
    # 结论
    ('审核结论', '人工结论',        'h_raw',         14, 'center'),
    ('审核结论', '人工结论(映射)',  'h_verdict',      9, 'center'),
    ('审核结论', 'AI结论',          'ai_result',      9, 'center'),
    ('审核结论', '建议赔付金额',    'payout',         9, 'center'),
    ('审核结论', '缺件清单',        'missing_str',   32, 'left'),
    ('审核结论', 'AI审核意见',      'remark',        38, 'left'),
    ('审核结论', '人工vs AI',       'match_str',      9, 'center'),
]

# 构建分组区间
groups_map = {}
for ci, (grp, *_) in enumerate(COLS, 1):
    groups_map.setdefault(grp, []).append(ci)

ROW_GROUP  = 1
ROW_HEADER = 2
ROW_DATA   = 3

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '行李延误审核明细'

# 分组标题
for grp, cols in groups_map.items():
    c1, c2 = cols[0], cols[-1]
    if c1 == c2:
        cell = ws.cell(ROW_GROUP, c1, grp)
    else:
        ws.merge_cells(start_row=ROW_GROUP, start_column=c1,
                       end_row=ROW_GROUP, end_column=c2)
        cell = ws.cell(ROW_GROUP, c1, grp)
    cell.fill      = hfill(C_GROUP_BG)
    cell.font      = hfont(C_GROUP_FG, bold=True, size=10)
    cell.alignment = halign('center', 'center')
    cell.border    = border

# 列标题
for ci, (grp, title, key, width, align) in enumerate(COLS, 1):
    cell = ws.cell(ROW_HEADER, ci, title)
    cell.fill      = hfill(C_HEADER_BG)
    cell.font      = hfont(C_HEADER_FG, bold=True, size=9)
    cell.alignment = halign('center', 'center', wrap=True)
    cell.border    = border
    ws.column_dimensions[get_column_letter(ci)].width = width

ws.row_dimensions[ROW_GROUP].height  = 20
ws.row_dimensions[ROW_HEADER].height = 34

# 数据行
for ri, row in enumerate(rows, ROW_DATA):
    is_odd = (ri - ROW_DATA) % 2 == 0
    base_bg = C_ODD if is_odd else C_EVEN

    for ci, (grp, title, key, width, align) in enumerate(COLS, 1):
        val = ri - ROW_DATA + 1 if key == '__idx__' else row.get(key, '')
        cell = ws.cell(ri, ci, val)
        cell.border    = border
        cell.alignment = halign(align, 'center', wrap=True)
        cell.font      = hfont(size=9)

        # 底色
        if title == '人工vs AI':
            vs = str(val)
            if '一致' in vs:    cell.fill = hfill(C_MATCH_OK)
            elif '不一致' in vs: cell.fill = hfill(C_MATCH_NO)
            else:               cell.fill = hfill(base_bg)
        elif title == 'AI结论':
            if '通过' in str(val):   cell.fill = hfill(C_MATCH_OK)
            elif '拒赔' in str(val): cell.fill = hfill(C_MATCH_NO)
            elif '补件' in str(val): cell.fill = hfill(C_SUPP)
            else:                   cell.fill = hfill(base_bg)
        elif title == '缺件清单' and val:
            cell.fill = hfill(C_SUPP)
        elif title in ('是否联程', '是否误机', '是否改签') and val == '是':
            cell.fill = hfill('EAF0FB')
        elif title in ('行李延误证明', '行李签收证明', '登机牌/机票') and val == '否':
            cell.fill = hfill(C_MATCH_NO)
        else:
            cell.fill = hfill(base_bg)

    ws.row_dimensions[ri].height = 30

ws.freeze_panes = f'A{ROW_DATA}'
ws.auto_filter.ref = (f'A{ROW_HEADER}:'
                      f'{get_column_letter(len(COLS))}{ROW_DATA+len(rows)-1}')


# ─── 汇总 Sheet ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet('汇总统计')

def s2r(ws2, r, label, val, bold=False, bg='D6E4F7'):
    c1 = ws2.cell(r, 1, label)
    c2 = ws2.cell(r, 2, val)
    for c in (c1, c2):
        c.font      = Font(name='微软雅黑', bold=bold, size=10)
        c.border    = border
        c.alignment = halign('left' if c.column == 1 else 'center', 'center')
    c1.fill = hfill(bg)
    c2.fill = hfill('FFFFFF')
    ws2.row_dimensions[r].height = 20

total     = len(rows)
n_app     = sum(1 for r in rows if r['ai_verdict'] == 'approve')
n_rej     = sum(1 for r in rows if r['ai_verdict'] == 'reject')
n_sup     = sum(1 for r in rows if r['ai_verdict'] == 'supplement')
n_unk     = sum(1 for r in rows if r['ai_verdict'] == 'unknown')
comp      = [r for r in rows if r['_h_verdict'] in ('approve', 'reject')]
match     = sum(1 for r in comp if r['_h_verdict'] == r['ai_verdict'])
acc       = match / len(comp) * 100 if comp else 0
n_conn    = sum(1 for r in rows if r['is_connecting'] == '是')
n_rebook  = sum(1 for r in rows if r['is_rebooking'] == '是')
n_missed  = sum(1 for r in rows if r['missed_conn'] == '是')

miss_cnt = Counter()
for r in rows:
    for m in (r['missing_str'].split('；') if r['missing_str'] else []):
        m = m.strip()
        if m:
            miss_cnt[m] += 1

mismatch_types = Counter()
for r in [r for r in comp if r['_h_verdict'] != r['ai_verdict']]:
    mismatch_types[f"人工{r['_h_verdict']}→AI {r['ai_verdict']}"] += 1

row = 1
title_cell = ws2.cell(row, 1, '行李延误 AI 审核汇总统计')
title_cell.font = Font(name='微软雅黑', bold=True, size=14, color='1F3864')
ws2.merge_cells(f'A{row}:B{row}')
ws2.row_dimensions[row].height = 28

row += 1; s2r(ws2, row, '── 总体情况 ──', '', bold=True, bg='1F3864')
ws2.cell(row, 1).font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
ws2.cell(row, 2).font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
row += 1; s2r(ws2, row, '总案件数', f'{total} 件', True)
row += 1; s2r(ws2, row, 'AI 审核通过', f'{n_app} 件 ({n_app/total*100:.1f}%)')
row += 1; s2r(ws2, row, 'AI 审核拒赔', f'{n_rej} 件 ({n_rej/total*100:.1f}%)')
row += 1; s2r(ws2, row, 'AI 需补件', f'{n_sup} 件 ({n_sup/total*100:.1f}%)')
row += 1; s2r(ws2, row, 'AI 未知', f'{n_unk} 件 ({n_unk/total*100:.1f}%)')

row += 1; s2r(ws2, row, '── 准确率分析 ──', '', bold=True, bg='1F3864')
ws2.cell(row, 1).font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
ws2.cell(row, 2).font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
row += 1; s2r(ws2, row, '可对比案件（人工有明确结论）', f'{len(comp)} 件', True)
row += 1; s2r(ws2, row, '结论一致', f'{match} 件 ({acc:.1f}%)', bg=C_MATCH_OK)
row += 1; s2r(ws2, row, '结论不一致', f'{len(comp)-match} 件 ({100-acc:.1f}%)', bg=C_MATCH_NO)
row += 1; s2r(ws2, row, '  其中：不一致类型分布', '', bold=True)
for k, v in sorted(mismatch_types.items(), key=lambda x: -x[1]):
    row += 1; s2r(ws2, row, f'    {k}', f'{v} 件')

row += 1; s2r(ws2, row, '── 行程特征 ──', '', bold=True, bg='1F3864')
ws2.cell(row, 1).font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
ws2.cell(row, 2).font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
row += 1; s2r(ws2, row, '联程案件', f'{n_conn} 件 ({n_conn/total*100:.1f}%)')
row += 1; s2r(ws2, row, '有改签情况', f'{n_rebook} 件 ({n_rebook/total*100:.1f}%)')
row += 1; s2r(ws2, row, '联程误机', f'{n_missed} 件')

row += 1; s2r(ws2, row, '── 缺件类型分布（Top 15）──', '', bold=True, bg='1F3864')
ws2.cell(row, 1).font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
ws2.cell(row, 2).font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
for miss, cnt in miss_cnt.most_common(15):
    row += 1
    s2r(ws2, row, f'  {miss}', f'{cnt} 件', bg=C_SUPP)

ws2.column_dimensions['A'].width = 38
ws2.column_dimensions['B'].width = 20

wb.save(OUT_FILE)
print(f'✅ Excel 已生成: {OUT_FILE}  ({len(rows)} 条记录)')
